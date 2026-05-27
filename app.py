import os
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    session,
    send_file,
    Response
)
from datetime import datetime
import pandas as pd
import joblib
import requests
import json
from sqlalchemy import text 
from utils import *
from pipeline.feature_engineering import preprocess_new_invoice
from pipeline.append_to_historical import append_historical_dataset
import plotly.express as px
import plotly
from database import engine
import plotly.graph_objects as go
import matplotlib
matplotlib.use('Agg')
import re
import matplotlib.pyplot as plt
from io import BytesIO
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)

# =========================================
# SECRET KEY
# =========================================
app.secret_key = 'payrisk_secret_key'
COMPANY_DB_MAP = {
    'dimas-jaya': {
        'db_name': 'PT DIMAS JAYA',
        'db_id': '2324081'
    },

    'safinah-laras': {
        'db_name': 'PT SAFINAH LARAS PERSADA',
        'db_id': None
    },

    'varia-karya': {
        'db_name': 'PT VARIA KARYA USAHA',
        'db_id': '2606316'
    },

    'mitra-besar': {
        'db_name': 'CV MITRA BESAR SEJATI',
        'db_id': '2089204'
    }
}
# =========================================
# ACCURATE CONFIG
# =========================================
CLIENT_ID = '83fba14d-6a2f-476e-a795-8b51aef39e3d'
CLIENT_SECRET = '3763bab66987b85a4c4341e2c8fc5366'

REDIRECT_URI = os.getenv(
    "ACCURATE_REDIRECT_URI"
)
TOKEN_FILE = 'token.json'
# =========================================
# LOAD MODEL & DATA
# =========================================
model = joblib.load('model/random_forest_model.pkl')
# =========================================
# LOAD TOKEN
# =========================================
def load_token():
    try:
        with open(TOKEN_FILE, 'r') as f:
            return json.load(f)
    except:
        return None


# =========================================
# SAVE TOKEN
# =========================================
def save_token(token_data):
    with open(TOKEN_FILE, 'w') as f:
        json.dump(token_data, f, indent=4)

# =========================================
# REFRESH ACCESS TOKEN
# =========================================
def refresh_access_token():
    token_data = load_token()
    if not token_data:
        return None
    refresh_token = token_data.get('refresh_token')
    response = requests.post(
        'https://account.accurate.id/oauth/token',
        data={
            'grant_type': 'refresh_token',
            'refresh_token': refresh_token
        },
        auth=(CLIENT_ID, CLIENT_SECRET)
    )
    new_token_data = response.json()
    print('REFRESH TOKEN RESPONSE:')
    print(new_token_data)

    # =====================================
    # JIKA BERHASIL REFRESH
    # =====================================
    if 'access_token' in new_token_data:
        save_token(new_token_data)
        session['access_token'] = new_token_data.get('access_token')
        session['refresh_token'] = new_token_data.get('refresh_token')
        return new_token_data.get('access_token')
    return None


# =========================================
# GET VALID ACCESS TOKEN
# =========================================
def get_valid_access_token():
    token_data = load_token()
    if not token_data:
        return None
    access_token = token_data.get('access_token')
    headers = {
        'Authorization': f'Bearer {access_token}'
    }
    try:

        # =====================================
        # TEST TOKEN
        # =====================================
        test_response = requests.get(
            'https://account.accurate.id/api/db-list.do',
            headers=headers,
            timeout=10
        )

        # =====================================
        # TOKEN EXPIRED
        # =====================================
        if test_response.status_code == 401:
            print('TOKEN EXPIRED -> REFRESHING')
            access_token = refresh_access_token()
        return access_token

    except requests.exceptions.Timeout:
        print("REQUEST TIMEOUT KE ACCURATE")
        return access_token

    except Exception as e:
        print("ERROR VALIDASI TOKEN:")
        print(e)

        return access_token

# =========================================
# HOME - SELECT COMPANY
# =========================================
@app.route('/')
def home():
    companies = [
        {
            'name': 'PT. DIMAS JAYA',
            'slug': 'dimas-jaya',
            'logo': 'dj.png'
        },

        {
            'name': 'PT. SAFINAH LARAS PERSADA',
            'slug': 'safinah-laras',
            'logo': 'slp.png'
        },

        {
            'name': 'PT. VARIA KARYA USAHA',
            'slug': 'varia-karya',
            'logo': 'vku.png'
        },

        {
            'name': 'CV. MITRA BESAR SEJATI',
            'slug': 'mitra-besar',
            'logo': 'mbs.png'
        }
    ]

    return render_template(
        'company_selector.html',
        companies=companies
    )

@app.route('/select-company/<company>')
def select_company(company):
    # =====================================
    # CEK LOGIN ACCURATE
    # =====================================
    access_token = get_valid_access_token()
    if not access_token:
        session['pending_company'] = company
        return redirect('/connect-accurate')
    # =====================================
    # SIMPAN COMPANY
    # =====================================
    session['company'] = company

    # =====================================
    # RESET DATABASE SESSION
    # =====================================
    session.pop('db_session', None)
    session.pop('db_host', None)
    session.pop('db_name', None)

    # =====================================
    # AMBIL DB ID
    # =====================================
    db_id = COMPANY_DB_MAP[company]['db_id']

    # =====================================
    # AUTO OPEN DATABASE
    # =====================================
    if db_id:
        headers = {
            'Authorization': f'Bearer {access_token}'
        }
        response = requests.get(
            f'https://account.accurate.id/api/open-db.do?id={db_id}',
            headers=headers
        )
        data = response.json()
        print("AUTO OPEN DB:")
        print(data)

        # =====================================
        # JIKA BERHASIL
        # =====================================
        if data.get('s') == True:
            session['db_session'] = data['session']
            session['db_host'] = data['host']
            session['db_name'] = COMPANY_DB_MAP[company]['db_name']
            print("DATABASE BERHASIL DIBUKA")
        else:
            print("GAGAL OPEN DATABASE")
            print(data)

    return redirect('/forecast')

# =========================================
# FORECASTING
# =========================================
@app.route('/forecast', methods=['GET', 'POST'])
def index():

    result = None
    customer_name = None
    if request.method == 'POST':
        df_hist = load_historical_data(
            session.get('company')
        )

        # =====================================
        # INPUT USER
        # =====================================
        customer_name = request.form['customer_name']
        customer_type = request.form['customer_type']
        order_type = request.form['order_type']
        amount = float(request.form['amount'])
        term = int(request.form['term'])

        # =====================================
        # PREPARE INPUT
        # =====================================
        input_df = prepare_input(
            customer_name,
            order_type,
            customer_type,
            amount,
            term,
            df_hist
        )

        # =====================================
        # PREDIKSI
        # =====================================
        status, prob = predict_payment_risk(
            input_df,
            model
        )

        # =====================================
        # ANALISIS TERM
        # =====================================
        df_terms = analyze_terms(
            customer_name,
            order_type,
            customer_type,
            amount,
            df_hist,
            model
        )

        # =====================================
        # REKOMENDASI
        # =====================================
        recommendation = recommend_term_business(prob)

        # =====================================
        # GENERATE PLOT
        # =====================================
        plot_term_risk(
            df_terms,
            rekom_term=recommendation['recommended_term']
        )

        # =====================================
        # DESKRIPSI
        # =====================================
        desc = generate_business_description(
            customer_name,
            df_hist,
            recommendation
        )

        # =====================================
        # SPECIAL NOTE
        # =====================================
        special_note = get_special_note(customer_name)

        # =====================================
        # RESULT
        # =====================================
        result = {
            'status': status,
            'prob': round(prob, 3),
            'prob_percent': round(prob * 100, 1),
            'recommendation': recommendation,
            'description': desc,
            'special_note': special_note
        }

        # =====================================
        # SIMPAN HISTORY FORECAST
        # =====================================
        print(session.get('company'))
        try:
            conn = engine.connect()
            insert_query = text("""
                INSERT INTO history_forecast (
                    customer_name,
                    customer_type,
                    order_type,
                    amount,
                    input_term,
                    company_name,
                    predicted_status,
                    probability,
                    recommended_term,
                    risk_category
                )
                VALUES (
                    :customer_name,
                    :customer_type,
                    :order_type,
                    :amount,
                    :input_term,
                    :company_name,
                    :predicted_status,
                    :probability,
                    :recommended_term,
                    :risk_category
                )
            """)

            conn.execute(insert_query, {
                "customer_name": customer_name,
                "customer_type": customer_type,
                "order_type": order_type,
                "amount": amount,
                "input_term": term,
                "predicted_status": status,
                "probability": float(prob),
                "recommended_term": recommendation['recommended_term'],
                "risk_category": recommendation['risk_category'],
                "company_name": session.get('company')
            })
            conn.commit()
            print("HISTORY FORECAST TERSIMPAN")
        except Exception as e:
            print("GAGAL SIMPAN HISTORY")
            print(e)
        finally:
            conn.close()

    return render_template(
        'index.html',
        result=result,
        customer_name=customer_name
    )

@app.route('/customer-suggestions')
def customer_suggestions():
    keyword = request.args.get(
        'q',
        ''
    )
    company = session.get('company')
    query = """
        SELECT DISTINCT "Customer"
        FROM sales_invoice_raw
        WHERE company_name = %(company)s
        AND LOWER("Customer") LIKE %(keyword)s
        ORDER BY "Customer"
        LIMIT 10
    """

    df = pd.read_sql(
        query,
        engine,
        params={
            'company': company,
            'keyword': f'%{keyword.lower()}%'
        }
    )

    suggestions = df[
        'Customer'
    ].tolist()
    return {
        'suggestions': suggestions
    }

# =========================================
# HALAMAN INTEGRASI ACCURATE
# =========================================
@app.route('/accurate')
def accurate_page():
    token_data = load_token()
    token = None
    if token_data:
        token = token_data.get('access_token')
    return render_template(
        'accurate.html',
        token=token,
        sync_logs=None
    )

# =========================================
# CONNECT ACCURATE
# =========================================
@app.route('/connect-accurate')
def connect_accurate():
    auth_url = (
        f'https://account.accurate.id/oauth/authorize?'
        f'client_id={CLIENT_ID}'
        f'&response_type=code'
        f'&redirect_uri={REDIRECT_URI}'
        f'&scope=customer_view sales_invoice_view'
    )
    return redirect(auth_url)

# =========================================
# CALLBACK OAUTH
# =========================================
@app.route('/callback')
def callback():
    code = request.args.get('code')
    token_url = 'https://account.accurate.id/oauth/token'
    response = requests.post(
        token_url,
        data={
            'code': code,
            'grant_type': 'authorization_code',
            'redirect_uri': REDIRECT_URI
        },
        auth=(CLIENT_ID, CLIENT_SECRET)
    )

    token_data = response.json()
    print(token_data)

    # =====================================
    # SIMPAN TOKEN
    # =====================================
    session['access_token'] = token_data.get('access_token')
    session['refresh_token'] = token_data.get('refresh_token')

    # DEBUG PATH
    print("CURRENT DIRECTORY:")
    print(os.getcwd())
    save_token(token_data)
    print("TOKEN BERHASIL DISIMPAN")

    pending_company = session.get('pending_company')
    if pending_company:
        return redirect(
            f'/select-company/{pending_company}'
        )
    return redirect('/accurate')

# # =========================================
# # OPEN DATABASE ACCURATE
# # =========================================
@app.route('/open-db/<db_id>')
def open_database(db_id):
    access_token = get_valid_access_token()

    headers = {
        'Authorization': f'Bearer {access_token}'
    }

    try:

        response = requests.get(
            f'https://account.accurate.id/api/open-db.do?id={db_id}',
            headers=headers,
            timeout=20
        )

        data = response.json()

        print("OPEN DB RESPONSE:")
        print(data)
        session['db_session'] = data['session']
        session['db_host'] = data['host']
        print("HOST TERSIMPAN:")
        print(session['db_host'])
        print("SESSION DB TERSIMPAN:")
        print(session['db_session'])

        return redirect('/accurate')

    except requests.exceptions.ConnectionError:
        return 'Gagal konek ke server Accurate'

    except requests.exceptions.Timeout:
        return 'Request timeout ke Accurate'

# =========================================
# CLEAN TERM
# =========================================
def clean_term(term_name):

    if not term_name:
        return 0

    term_name = str(term_name).upper()

    # COD
    if 'COD' in term_name or 'C.O.D' in term_name:
        return 0

    # AMBIL ANGKA
    match = re.search(r'(\d+)', term_name)
    if match:
        return int(match.group(1))
    return 0

# =========================================
# GET SALES INVOICE -> EXPORT EXCEL
# =========================================
def get_invoices(
    company=None,
    sync_mode=None,
    start_date=None,
    end_date=None
):

    host = session.get('db_host')
    db_session = session.get('db_session')
    access_token = get_valid_access_token()

    if not host:
        return 'Database belum dibuka'

    headers = {
        'Authorization': f'Bearer {access_token}',
        'X-Session-ID': db_session
    }

    hasil = []
    page = 1
    total_pages = 1

    print('AMBIL DATA INVOICE...')

    while page <= total_pages:

        print(f'\nAMBIL PAGE {page}')
        url = (
            f'{host}/accurate/api/sales-invoice/list.do'
            f'?sp.page={page}'
            f'&sp.pageSize=100'
            f'&fields=id,number,customer,status,transDate,dueDate,paymentTerm,totalAmount,paidAmount'
        )

        if start_date and end_date:
            start_date = datetime.strptime(
                start_date,
                '%Y-%m-%d'
            ).strftime('%d/%m/%Y')
            end_date = datetime.strptime(
                end_date,
                '%Y-%m-%d'
            ).strftime('%d/%m/%Y')
            url += (
                f'&filter.transDate.op=BETWEEN'
                f'&filter.transDate.val[0]={start_date}'
                f'&filter.transDate.val[1]={end_date}'
            )
        response = requests.get(
            url,
            headers=headers
        )
        response_json = response.json()
        data = response_json.get('d', [])

        # =========================
        # AMBIL TOTAL PAGE
        # =========================
        sp = response_json.get('sp', {})
        total_pages = sp.get('pageCount', 1)
        print(response_json.get('sp'))
        print(f'Total invoice page {page}: {len(data)}')
        print(f'Total pages: {total_pages}')

        # =========================
        # LOOP INVOICE
        # =========================
        for inv in data:
            invoice_id = inv.get('id')

            try:
                detail_response = requests.get(
                    f'{host}/accurate/api/sales-invoice/detail.do?id={invoice_id}',
                    headers=headers
                )

                detail = detail_response.json()['d']

                payment_date = detail.get('lastPaymentDate')
                due_date = detail.get('dueDate')

                receipt_history = detail.get(
                    'receiptHistory',
                    []
                )

                payment_method = None

                if receipt_history:
                    payment_method = receipt_history[0].get(
                        'historyPaymentName'
                    )

                total_amount = detail.get(
                    'totalAmount',
                    0
                )

                balance_due = (
                    total_amount
                    if detail.get('statusName') != 'Lunas'
                    else 0
                )

                # =====================================
                # HITUNG DELAY
                # =====================================
                due_date_obj = datetime.strptime(
                    due_date,
                    '%d/%m/%Y'
                ).date()

                today = datetime.today().date()
                delay = 0
                # =====================================
                # SUDAH LUNAS
                # =====================================
                if detail.get('statusName') == 'Lunas':
                    if payment_date:
                        payment_date_obj = datetime.strptime(
                            payment_date,
                            '%d/%m/%Y'
                        ).date()
                        delay = (
                            payment_date_obj - due_date_obj
                        ).days

                # =====================================
                # BELUM LUNAS
                # =====================================
                else:
                    delay = (
                        today - due_date_obj
                    ).days

                hasil.append({
                    'source_system': 'ACCURATE',
                    'company_name': session.get('company'),
                    'Transaction ID': detail.get('id'),
                    'Invoice Number': detail.get('number'),
                    'Customer': detail.get(
                        'customer',
                        {}
                    ).get('name'),
                    'Status': detail.get('statusName'),
                    'Transaction Date': detail.get('transDate'),
                    'Due Date': due_date,
                    'Term': clean_term(
                        detail.get(
                            'paymentTerm',
                            {}
                        ).get('name')
                    ),
                    'Payment Date': payment_date,
                    'Payment Method': payment_method,
                    'Total Amount': total_amount,
                    'Balance Due': balance_due,
                    'Delay': delay,
                })

                print(
                    f"Berhasil ambil invoice: {detail.get('number')}"
                )
            except Exception as e:
                print(f'ERROR invoice {invoice_id}')
                print(e)

        # =========================
        # NEXT PAGE
        # =========================
        page += 1
    print(f'\nTOTAL FINAL INVOICE: {len(hasil)}')

    # =========================
    # EXPORT EXCEL
    # =========================
    df = pd.DataFrame(hasil)
    df['Transaction Date'] = pd.to_datetime(
        df['Transaction Date'],
        format='%d/%m/%Y',
        errors='coerce'
    )
    df['Due Date'] = pd.to_datetime(
        df['Due Date'],
        format='%d/%m/%Y',
        errors='coerce'
    )
    df['Payment Date'] = pd.to_datetime(
        df['Payment Date'],
        format='%d/%m/%Y',
        errors='coerce'
    )
    numeric_cols = [
        'Total Amount',
        'Balance Due',
        'Delay'
    ]

    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    conn = engine.connect()
    for _, row in df.iterrows():
        payment_date = row['Payment Date']
        payment_method = row['Payment Method']
        delay = row['Delay']

        # =========================
        # HANDLE NaT / NaN
        # =========================
        if pd.isna(payment_date):
            payment_date = None

        if pd.isna(payment_method):
            payment_method = None

        if pd.isna(delay):
            delay = None

        check_query = text("""
            SELECT COUNT(*)
            FROM sales_invoice_raw
            WHERE source_system = :source_system
            AND company_name = :company_name
            AND "Transaction ID" = :trx_id
        """)

        result = conn.execute(
            check_query,
            {
                "source_system": row['source_system'],
                "trx_id": row['Transaction ID'],
                "company_name": row['company_name']
            }
        )
        exists = result.scalar()

        # ====================================
        # JIKA BELUM ADA -> INSERT
        # ====================================
        if exists == 0:
            insert_query = text("""
                INSERT INTO sales_invoice_raw (
                    company_name,
                    source_system,
                    "Transaction ID",
                    "Invoice Number",
                    "Customer",
                    "Status",
                    "Transaction Date",
                    "Due Date",
                    "Term",
                    "Payment Date",
                    "Payment Method",
                    "Total Amount",
                    "Balance Due",
                    "Delay"
                )
                VALUES (
                    :company_name,
                    :source_system,
                    :transaction_id,
                    :invoice_number,
                    :customer,
                    :status,
                    :transaction_date,
                    :due_date,
                    :term,
                    :payment_date,
                    :payment_method,
                    :total_amount,
                    :balance_due,
                    :delay
                )
            """)

            conn.execute(insert_query, {
                "source_system": row['source_system'],
                "company_name": row['company_name'],
                "transaction_id": row['Transaction ID'],
                "invoice_number": row['Invoice Number'],
                "customer": row['Customer'],
                "status": row['Status'],
                "transaction_date": row['Transaction Date'],
                "due_date": row['Due Date'],
                "term": row['Term'],
                "payment_date": payment_date,
                "payment_method": payment_method,
                "total_amount": row['Total Amount'],
                "balance_due": row['Balance Due'],
                "delay": delay
            })
            print(f"INSERT: {row['Invoice Number']}")

        # ====================================
        # JIKA SUDAH ADA -> UPDATE
        # ====================================
        else:
            update_query = text("""
                UPDATE sales_invoice_raw
                SET
                    "Status" = :status,
                    "Payment Date" = :payment_date,
                    "Payment Method" = :payment_method,
                    "Balance Due" = :balance_due,
                    "Delay" = :delay
                WHERE source_system = :source_system
                AND company_name = :company_name
                AND "Transaction ID" = :transaction_id
            """)

            conn.execute(update_query, {
                "source_system": row['source_system'],
                "transaction_id": row['Transaction ID'],
                "company_name": row['company_name'],
                "status": row['Status'],
                "payment_date": payment_date,
                "payment_method": payment_method,
                "balance_due": row['Balance Due'],
                "delay": delay
            })

            print(f"UPDATE: {row['Invoice Number']}")

    conn.commit()
    conn.close()

    print('DATABASE UPDATED')
    return 'DATABASE UPDATED'


@app.route('/sync-data', methods=['POST'])
def sync_data():
    company = session.get('company')
    sync_mode = request.form.get('sync_mode')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    sync_logs = []
    token_data = load_token()
    token = None

    if token_data:
        token = token_data.get('access_token')

    # =====================================
    # SYNC DATA
    # =====================================
    sync_logs.append('Mengambil data invoice...')

    get_invoices(
        company=company,
        sync_mode=sync_mode,
        start_date=start_date,
        end_date=end_date
    )
    sync_logs.append('Preprocessing data invoice...')
    preprocess_new_invoice(company)
    sync_logs.append('Update historical dataset...')
    append_historical_dataset(company)
    sync_logs.append('Sinkronisasi selesai!')
    return render_template(
        'accurate.html',
        token=token,
        sync_logs=sync_logs
    )
# =========================================
# HISTORY FORECAST
# =========================================
@app.route('/history')
def history():
    company = session.get('company')
    query = """
        SELECT *
        FROM history_forecast
        WHERE company_name = %(company)s
        ORDER BY id DESC
    """

    df = pd.read_sql(
        query,
        engine,
        params={
            'company': company
        }
    )

    history_data = df.to_dict(orient='records')

    return render_template(
        'history.html',
        history_data=history_data
    )

# =========================================
# DASHBOARD
# =========================================
@app.route('/dashboard')
def dashboard():

    # =====================================
    # LOAD DATA
    # =====================================
    company = session.get('company')
    df_hist = pd.read_sql(
        """
        SELECT *
        FROM historical_invoice
        WHERE company_name = %(company)s
        """,
        engine,
        params={
            'company': company
        }
    )

    df_forecast = pd.read_sql(
        """
        SELECT *
        FROM history_forecast
        WHERE company_name = %(company)s
        """,
        engine,
        params={
            'company': company
        }
    )

    # =====================================
    # KPI
    # =====================================
    total_invoice = len(df_hist)
    total_customer = df_hist[
        'Customer'
    ].nunique()

    avg_delay = round(
        df_hist['Delay'].mean(),
        2
    )

    high_risk = len(
        df_hist[
            df_hist['is_late'] == 1
        ]
    )

    # =====================================
    # NORMALISASI IS_LATE
    # =====================================
    df_hist['is_late'] = (
        df_hist['is_late']
        .astype(str)
        .str.strip()
        .str.lower()
    )

    df_hist['is_late'] = df_hist['is_late'].map({
        '1': 1,
        '0': 0,
        'late': 1,
        'on time': 0,
        'yes': 1,
        'no': 0,
        'true': 1,
        'false': 0
    })
    # =====================================
    # TERM CLEAN
    # =====================================
    df_hist['Term Clean'] = (
        df_hist['Term']
        .astype(str)
        .str.extract(r'(\d+)', expand=False)
    )

    df_hist['Term Clean'] = pd.to_numeric(
        df_hist['Term Clean'],
        errors='coerce'
    ).fillna(0)

    # =====================================
    # TERM CATEGORY
    # =====================================
    def categorize_term(x):
        if x == 0:
            return '0 Hari'
        elif x <= 14:
            return '<=14 Hari'
        elif x <= 30:
            return '<=30 Hari'
        else:
            return '>30 Hari'
    df_hist['Term Category'] = (
        df_hist['Term Clean']
        .apply(categorize_term)
    )

    # =====================================
    # CUSTOMER RATIO
    # =====================================
    customer_ratio = (
        df_hist
        .groupby('Kategori Customer')['is_late']
        .mean()
        * 100
    ).round(1)
    customer_labels = customer_ratio.index.tolist()
    customer_values = customer_ratio.values.tolist()

    # =====================================
    # TERM RATIO
    # =====================================
    term_ratio = (
        df_hist
        .groupby('Term Category')['is_late']
        .mean()
        * 100
    ).round(1)
    term_labels = term_ratio.index.tolist()
    term_values = term_ratio.values.tolist()
    # =====================================
    # TOP 5 CUSTOMER
    # =====================================
    top_customer = (
        df_hist
        .groupby('Customer')
        .agg(
            average_delay=('Delay', 'mean'),
            total_invoice=('Delay', 'count'),
            max_delay=('Delay', 'max')
        )

        .sort_values(
            by='average_delay',
            ascending=False
        )
        .head(5)
        .round(1)
        .reset_index()
    )

    top_labels = top_customer['Customer'].tolist()
    top_chart_data = []
    for _, row in top_customer.iterrows():
        top_chart_data.append({
            'value': float(row['average_delay']),
            'total': int(row['total_invoice']),
            'max_delay': float(row['max_delay'])
        })
    # =====================================
    # CHART 4
    # HISTORICAL PAYMENT STATUS
    # =====================================
    df_hist['is_late'] = pd.to_numeric(
        df_hist['is_late'],
        errors='coerce'
    )
    historical_chart = pd.DataFrame({
        'Status': ['Late', 'On Time'],
        'Count': [
            int((df_hist['is_late'] == 1).sum()),
            int((df_hist['is_late'] == 0).sum())
        ]
    })

    historical_chart['Count'] = pd.to_numeric(
        historical_chart['Count']
    )

    # =====================================
    # DONUT CHART
    # =====================================
    late_count = int((df_hist['is_late'] == 1).sum())
    ontime_count = int((df_hist['is_late'] == 0).sum())
    fig_historical = go.Figure(
        data=[
            go.Pie(
                labels=['Late', 'On Time'],
                values=[late_count, ontime_count],
                hole=0.6,
                textinfo='none',
                texttemplate='%{percent}',
                textposition='inside',
                marker=dict(
                    colors=[
                        '#ef4444',  # Late merah
                        '#3b82f6'   # On Time biru
                    ]
                )
            )
        ]
    )

    fig_historical.update_layout(
        title='Historical Payment Status Distribution'
    )

    # =====================================
    # STYLE
    # =====================================
    fig_historical.update_traces(
        textposition='inside',
        textinfo='label+value+percent'
    )

    historical_graph = json.dumps(
        fig_historical,
        cls=plotly.utils.PlotlyJSONEncoder
    )

    # =====================================
    # MONTHLY TREND
    # =====================================
    df_hist['Transaction Date'] = pd.to_datetime(
        df_hist['Transaction Date']
    )

    # =====================================
    # FILTER DELAY
    # =====================================
    df_delay_only = df_hist[
        df_hist['Delay'] > 0
    ]

    # =====================================
    # GROUP MONTHLY
    # =====================================
    monthly = (
        df_delay_only
        .groupby(
            df_delay_only['Transaction Date']
            .dt.to_period('M')
        )['Delay']
        .mean()
        .reset_index()
    )

    # =====================================
    # CONVERT
    # =====================================
    monthly['Transaction Date'] = (
        monthly['Transaction Date']
        .astype(str)
    )
    monthly = monthly.tail(12)
    print(monthly)

    # =====================================
    # MANUAL LINE CHART
    # =====================================
    fig_month = go.Figure(
        data=[
            go.Scatter(
                x=monthly['Transaction Date'].tolist(),
                y=monthly['Delay'].tolist(),
                mode='lines+markers',
                line=dict(
                    width=4
                ),
                marker=dict(
                    size=10
                )
            )
        ]
    )

    # =====================================
    # LAYOUT
    # =====================================
    fig_month.update_layout(
        title='Monthly Delay Trend (Last 12 Months)',
        xaxis_title='Month',
        yaxis_title='Average Delay (Days)',
        yaxis=dict(
            range=[0,25]
        )
    )
    print(fig_month)
    month_graph = json.dumps(
        fig_month,
        cls=plotly.utils.PlotlyJSONEncoder
    )

    # =====================================
    # CUSTOMER DETAIL
    # =====================================
    customer_detail = (
        df_hist
        .groupby('Kategori Customer')
        .agg(
            total_invoice=('is_late', 'count'),
            invoice_delay=('is_late', 'sum')
        )
    )

    customer_percentage = (
        df_hist
        .groupby('Kategori Customer')['is_late']
        .mean()
        * 100
    ).round(1)

    customer_chart_data = []
    for idx in customer_detail.index:
        customer_chart_data.append({
            'value': float(customer_percentage[idx]),
            'total': int(customer_detail.loc[idx, 'total_invoice']),
            'late': int(customer_detail.loc[idx, 'invoice_delay'])
        })

    # =====================================
    # TERM DETAIL
    # =====================================
    term_detail = (
        df_hist
        .groupby('Term Category')
        .agg(
            total_invoice=('is_late', 'count'),
            invoice_delay=('is_late', 'sum')
        )
    )

    term_percentage = (
        df_hist
        .groupby('Term Category')['is_late']
        .mean()
        * 100
    ).round(1)

    term_chart_data = []
    for idx in term_detail.index:
        term_chart_data.append({
            'value': float(term_percentage[idx]),
            'total': int(term_detail.loc[idx, 'total_invoice']),
            'late': int(term_detail.loc[idx, 'invoice_delay'])
        })
    return render_template(
        'dashboard.html',
        total_invoice=total_invoice,
        total_customer=total_customer,
        avg_delay=avg_delay,
        high_risk=high_risk,
        historical_graph=historical_graph,
        month_graph=month_graph,
        customer_labels=customer_labels,
        customer_values=customer_values,
        term_labels=term_labels,
        term_values=term_values,
        customer_chart_data=customer_chart_data,
        term_chart_data=term_chart_data,
        top_labels=top_labels,
        top_chart_data=top_chart_data,
    )

# =========================================
# GET DATABASE LIST
# =========================================
@app.route('/update-data')
def update_data():
    access_token = get_valid_access_token()
    if not access_token:
        return 'Belum connect Accurate'
    db_data = get_database_list(
        access_token
    )
    return db_data

@app.route('/company-selector')
def company_selector():
    companies = [
        {
            'name': 'PT. DIMAS JAYA',
            'slug': 'dimas-jaya',
            'logo': 'dj.png'
        },
        {
            'name': 'PT. SAFINAH LARAS PERSADA',
            'slug': 'safinah-laras',
            'logo': 'slp.png'
        },
        {
            'name': 'PT. VARIA KARYA USAHA',
            'slug': 'varia-karya',
            'logo': 'vku.png'
        },
        {
            'name': 'CV. MITRA BESAR SEJATI',
            'slug': 'mitra-besar',
            'logo': 'mbs.png'
        }
    ]

    return render_template(
        'company_selector.html',
        companies=companies
    )

@app.route('/logout')
def logout():
    # =====================================
    # HAPUS SESSION
    # =====================================
    session.clear()
    # =====================================
    # HAPUS TOKEN
    # =====================================
    if os.path.exists('token.json'):
        os.remove('token.json')
        print('TOKEN DIHAPUS')
    return redirect('/company-selector')


################# KERJAAN SETELAH DIBERI MASUKAN ####################################################################
@app.route('/customer/<customer_name>')
def customer_summary(customer_name):
    company = session.get('company')
    query = """
        SELECT *
        FROM sales_invoice_raw
        WHERE "Customer" = %(customer)s
        AND company_name = %(company)s
        ORDER BY "Transaction Date" DESC
    """

    df = pd.read_sql(
        query,
        engine,
        params={
            'customer': customer_name,
            'company': company
        }
    )

    # =====================================
    # KPI
    # =====================================
    total_invoice = len(df)
    total_amount = df[
        'Total Amount'
    ].sum()

    # =====================================
    # FIRST & LAST TRANSACTION
    # =====================================
    first_transaction_date = pd.to_datetime(
        df['Transaction Date'],
        dayfirst=True,
        errors='coerce'
    ).min()

    last_transaction_date = pd.to_datetime(
        df['Transaction Date'],
        dayfirst=True,
        errors='coerce'
    ).max()

    # =====================================
    # FORMAT DATE
    # =====================================
    if pd.notnull(first_transaction_date):
        first_transaction = first_transaction_date.strftime(
            '%d-%m-%Y'
        )
    else:
        first_transaction = '-'
    
    if pd.notnull(last_transaction_date):
        last_transaction = last_transaction_date.strftime(
            '%d-%m-%Y'
        )
    else:
        last_transaction = '-'

    avg_delay = round(
        df['Delay'].mean(),
        2
    )

    unpaid_df = df[
        df['Status'] != 'Lunas'
    ]

    paid_df = df[
        df['Status'] == 'Lunas'
    ]

    unpaid_count = len(unpaid_df)
    outstanding_amount = unpaid_df[
        'Balance Due'
    ].sum()

    paid_invoice = len(paid_df)
    paid_amount = paid_df[
        'Total Amount'
    ].sum()

    # =====================================
    # FORMAT DATE
    # =====================================
    date_cols = [
        'Transaction Date',
        'Due Date',
        'Payment Date'
    ]

    for col in date_cols:
        df[col] = pd.to_datetime(
            df[col],
            errors='coerce'
        ).dt.strftime('%d-%m-%Y')

    # =====================================
    # RECREATE TABLE
    # =====================================
    unpaid_df = df[
        df['Status'] != 'Lunas'
    ]
    paid_df = df[
        df['Status'] == 'Lunas'
    ]
    outstanding_table = unpaid_df.to_dict(
        orient='records'
    )
    payment_table = paid_df.to_dict(
        orient='records'
    )
    # =====================================
    # LOAD NOTES
    # =====================================
    notes_query = """
        SELECT *
        FROM customer_notes
        WHERE customer_name = %(customer)s
        AND company_name = %(company)s
        ORDER BY created_at DESC
    """
    notes_df = pd.read_sql(
        notes_query,
        engine,
        params={
            'customer': customer_name,
            'company': company
        }
    )

    notes_df['created_at'] = pd.to_datetime(
        notes_df['created_at']
    ).dt.strftime('%d-%m-%Y %H:%M')

    notes_data = notes_df.to_dict(
        orient='records'
    )

    # =====================================
    # NOTES
    # =====================================
    behavior_notes = generate_customer_behavior(df)
    
    return render_template(
        'customer_summary.html',
        customer_name=customer_name,
        tables=df.to_dict(orient='records'),
        outstanding_table=outstanding_table,
        behavior_notes=behavior_notes,
        total_invoice=total_invoice,
        total_amount=total_amount,
        avg_delay=avg_delay,
        unpaid_count=unpaid_count,
        outstanding_amount=outstanding_amount,
        paid_amount=paid_amount,  
        payment_table=payment_table,
        paid_invoice=paid_invoice,
        first_transaction=first_transaction,
        notes_data=notes_data,
        last_transaction=last_transaction
    )

@app.route('/save-note', methods=['POST'])
def save_note():
    company = session.get('company')
    customer_name = request.form['customer_name']
    note = request.form['note']
    conn = engine.connect()
    insert_query = text("""
        INSERT INTO customer_notes (
            company_name,
            customer_name,
            note
        )
        VALUES (
            :company_name,
            :customer_name,
            :note
        )
    """)
    conn.execute(insert_query, {
        'company_name': company,
        'customer_name': customer_name,
        'note': note
    })

    conn.commit()
    conn.close()

    return redirect(
        f'/customer/{customer_name}'
    )

@app.route('/delete-note/<int:note_id>/<customer_name>')
def delete_note(note_id, customer_name): 
    conn = engine.connect()
    delete_query = text("""
        DELETE FROM customer_notes
        WHERE id = :id
    """)

    conn.execute(delete_query, {
        'id': note_id
    })

    conn.commit()
    conn.close()

    return redirect(
        f'/customer/{customer_name}'
    )

@app.route('/edit-note', methods=['POST'])
def edit_note():
    note_id = request.form['note_id']
    customer_name = request.form['customer_name']
    updated_note = request.form['updated_note']
    conn = engine.connect()
    update_query = text("""
        UPDATE customer_notes
        SET note = :note
        WHERE id = :id
    """)

    conn.execute(update_query, {
        'note': updated_note,
        'id': note_id
    })

    conn.commit()
    conn.close()

    return redirect(
        f'/customer/{customer_name}'
    )

@app.route('/download-customer-excel/<customer_name>')
def download_customer_excel(customer_name):
    # =====================================
    # GET COMPANY
    # =====================================
    company = session.get('company')
    # =====================================
    # QUERY
    # =====================================

    query = """
        SELECT *
        FROM sales_invoice_raw
        WHERE "Customer" = %(customer)s
        AND company_name = %(company)s
        ORDER BY "Transaction Date" DESC
    """
    # =====================================
    # LOAD DATA
    # =====================================
    df = pd.read_sql(
        query,
        engine,
        params={
            'customer': customer_name,
            'company': company
        }
    )
    # =====================================
    # KPI
    # =====================================
    total_invoice = len(df)
    total_amount = df[
        'Total Amount'
    ].sum()
    # =====================================
    # FIRST & LAST TRANSACTION
    # =====================================
    first_transaction_date = pd.to_datetime(
        df['Transaction Date'],
        dayfirst=True,
        errors='coerce'
    ).min()
    last_transaction_date = pd.to_datetime(
        df['Transaction Date'],
        dayfirst=True,
        errors='coerce'
    ).max()

    # =====================================
    # FORMAT DATE
    # =====================================
    if pd.notnull(first_transaction_date):
        first_transaction = first_transaction_date.strftime(
            '%d-%m-%Y'
        )
    else:
        first_transaction = '-'
    if pd.notnull(last_transaction_date):
        last_transaction = last_transaction_date.strftime(
            '%d-%m-%Y'
        )
    else:
        last_transaction = '-'
    avg_delay = round(
        df['Delay'].mean(),
        2
    )
    unpaid_df = df[
        df['Status'] != 'Lunas'
    ]
    paid_df = df[
        df['Status'] == 'Lunas'
    ]

    # =====================================
    # FORMAT DATE
    # =====================================
    date_cols = [
        'Transaction Date',
        'Due Date',
        'Payment Date'
    ]

    for col in date_cols:
        df[col] = pd.to_datetime(
            df[col],
            errors='coerce'
        ).dt.strftime('%d-%m-%Y')
    # =====================================
    # SPLIT DATA
    # =====================================
    unpaid_df = df[
        df['Status'] != 'Lunas'
    ]
    paid_df = df[
        df['Status'] == 'Lunas'
    ]
    # =====================================
    # KPI
    # =====================================
    total_invoice = len(df)
    total_transaction = df[
        'Total Amount'
    ].sum()
    avg_delay = round(
        df['Delay'].mean(),
        2
    )
    paid_invoice = len(paid_df)
    paid_amount = paid_df[
        'Total Amount'
    ].sum()
    unpaid_count = len(unpaid_df)
    outstanding_amount = unpaid_df[
        'Balance Due'
    ].sum()
    # =====================================
    # BEHAVIOR
    # =====================================
    behavior_notes = generate_customer_behavior(df)
    # =====================================
    # COLLECTION NOTES
    # =====================================
    notes_query = text("""
        SELECT *
        FROM customer_notes
        WHERE customer_name = :customer
        ORDER BY created_at DESC
    """)
    notes_data = pd.read_sql(
        notes_query,
        engine,
        params={
            'customer': customer_name,
            'company': company
        }
    )

    # =====================================
    # CREATE WORKBOOK
    # =====================================
    wb = Workbook()
    ws = wb.active
    ws.title = 'Customer Analysis'

    # =====================================
    # STYLE
    # =====================================
    title_fill = PatternFill(
        "solid",
        fgColor="D4A017"
    )
    blue_header_fill = PatternFill(
        "solid",
        fgColor="1D4ED8"
    )
    green_fill = PatternFill(
        "solid",
        fgColor="DCFCE7"
    )
    red_fill = PatternFill(
        "solid",
        fgColor="FEE2E2"
    )
    yellow_fill = PatternFill(
        "solid",
        fgColor="DBEAFE"
    )
    thin = Side(
        border_style="thin",
        color="D1D5DB"
    )
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )
    # =====================================
    # TITLE
    # =====================================
    ws.merge_cells('B2:H3')
    title_cell = ws['B2']
    title_cell.value = f'{customer_name}'
    title_cell.font = Font(
        size=24,
        bold=True,
        color='FFFFFF'
    )
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # =====================================
    # GENERAL OVERVIEW
    # =====================================
    ws.merge_cells('B5:C5')
    ws['B5'] = 'GENERAL OVERVIEW'
    ws['B5'].font = Font(
        bold=True,
        size=14,
        color='FFFFFF'
    )
    ws['B5'].fill = blue_header_fill

    general_data = [
        ['Total Invoice', total_invoice],
        ['Total Transaction', f'Rp {total_transaction:,.0f}'],
        ['Average Delay', f'{avg_delay} hari'],
        ['First Transaction', first_transaction],
        ['Last Transaction', last_transaction]
    ]
    row = 6

    for item in general_data:
        ws[f'B{row}'] = item[0]
        ws[f'C{row}'] = item[1]
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = yellow_fill
        ws[f'B{row}'].border = border
        ws[f'C{row}'].border = border
        row += 1

    # =====================================
    # PAYMENT SUCCESS
    # =====================================
    ws.merge_cells('E5:F5')
    ws['E5'] = 'PAYMENT SUCCESS'
    ws['E5'].font = Font(
        bold=True,
        size=14,
        color='FFFFFF'
    )
    ws['E5'].fill = PatternFill(
        "solid",
        fgColor="16A34A"
    )

    payment_data = [
        ['Paid Invoice', paid_invoice],
        ['Payment Received', f'Rp {paid_amount:,.0f}']
    ]
    row2 = 6

    for item in payment_data:
        ws[f'E{row2}'] = item[0]
        ws[f'F{row2}'] = item[1]
        ws[f'E{row2}'].font = Font(bold=True)
        ws[f'E{row2}'].fill = green_fill
        ws[f'E{row2}'].border = border
        ws[f'F{row2}'].border = border
        row2 += 1

    # =====================================
    # OUTSTANDING RISK
    # =====================================
    ws.merge_cells('H5:I5')
    ws['H5'] = 'OUTSTANDING RISK'
    ws['H5'].font = Font(
        bold=True,
        size=14,
        color='FFFFFF'
    )

    ws['H5'].fill = PatternFill(
        "solid",
        fgColor="DC2626"
    )

    outstanding_data = [
        ['Outstanding Invoice', unpaid_count],
        ['Outstanding Amount', f'Rp {outstanding_amount:,.0f}']
    ]
    row3 = 6

    for item in outstanding_data:
        ws[f'H{row3}'] = item[0]
        ws[f'I{row3}'] = item[1]
        ws[f'H{row3}'].font = Font(bold=True)
        ws[f'H{row3}'].fill = red_fill
        ws[f'H{row3}'].border = border
        ws[f'I{row3}'].border = border
        row3 += 1
    # =====================================
    # NEXT ROW
    # =====================================
    row = 13

    # =====================================
    # BEHAVIOR ANALYSIS
    # =====================================

    row += 2
    ws[f'B{row}'] = 'BEHAVIOR ANALYSIS'
    ws[f'B{row}'].font = Font(
        bold=True,
        size=16
    )
    row += 2

    for note in behavior_notes:
        ws[f'B{row}'] = f'• {note}'
        row += 1

    # =====================================
    # COLLECTION INSIGHTS
    # =====================================

    row += 2
    ws[f'B{row}'] = 'COLLECTION INSIGHTS'
    ws[f'B{row}'].font = Font(
        bold=True,
        size=16
    )
    row += 2

    if len(notes_data) > 0:
        for _, note in notes_data.iterrows():
            ws[f'B{row}'] = str(
                note['created_at']
            )

            row += 1
            ws[f'B{row}'] = note['note']
            ws[f'B{row}'].alignment = Alignment(
                wrap_text=True
            )
            ws[f'B{row}'].fill = green_fill
            ws[f'B{row}'].border = border
            row += 2
    else:
        ws[f'B{row}'] = 'No collection notes.'

    # =====================================
    # OUTSTANDING TABLE
    # =====================================
    row += 2
    ws[f'B{row}'] = 'OUTSTANDING INVOICE'
    ws[f'B{row}'].font = Font(
        bold=True,
        size=16
    )
    row += 2

    outstanding_headers = [
        'Invoice Number',
        'Transaction Date',
        'Due Date',
        'Payment Date',
        'Total Amount',
        'Balance Due',
        'Delay'
    ]

    for col_num, header in enumerate(
        outstanding_headers,
        2
    ):

        cell = ws.cell(
            row=row,
            column=col_num
        )
        cell.value = header
        cell.font = Font(
            bold=True
        )
        cell.fill = red_fill
        cell.border = border
    row += 1

    for _, data in unpaid_df.iterrows():
        values = [
            data['Invoice Number'],
            data['Transaction Date'],
            data['Due Date'],
            data['Payment Date'],
            f"Rp {data['Total Amount']:,.0f}",
            f"Rp {data['Balance Due']:,.0f}",
            f"{int(data['Delay'])} hari"
            if pd.notnull(data['Delay'])
            else "-"
        ]

        for col_num, value in enumerate(
            values,
            2
        ):

            cell = ws.cell(
                row=row,
                column=col_num
            )
            cell.value = value
            cell.border = border
        row += 1
    # =====================================
    # PAYMENT SUCCESS TABLE
    # =====================================
    row += 3
    ws[f'B{row}'] = 'PAYMENT SUCCESS HISTORY'
    ws[f'B{row}'].font = Font(
        bold=True,
        size=16
    )
    row += 2

    payment_headers = [
        'Invoice Number',
        'Transaction Date',
        'Due Date',
        'Payment Date',
        'Total Amount',
        'Delay'
    ]

    for col_num, header in enumerate(
        payment_headers,
        2
    ):

        cell = ws.cell(
            row=row,
            column=col_num
        )

        cell.value = header
        cell.font = Font(
            bold=True
        )
        cell.fill = green_fill
        cell.border = border
    row += 1

    for _, data in paid_df.iterrows():
        values = [
            data['Invoice Number'],
            data['Transaction Date'],
            data['Due Date'],
            data['Payment Date'],
            f"Rp {data['Total Amount']:,.0f}",
            f"{int(data['Delay'])} hari"
            if pd.notnull(data['Delay'])
            else "-"
        ]

        for col_num, value in enumerate(
            values,
            2
        ):
            cell = ws.cell(
                row=row,
                column=col_num
            )
            cell.value = value
            cell.border = border
        row += 1

    # =====================================
    # AUTO WIDTH
    # =====================================

    for column_cells in ws.columns:
        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            get_column_letter(
                column_cells[0].column
            )
        ].width = length + 5

    # =====================================
    # SAVE
    # =====================================

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'{customer_name}_analysis.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# @app.route('/download-customer-pdf/<customer_name>')
# def download_customer_pdf(customer_name):
#     url = request.host_url + f'customer/{customer_name}?pdf=1'
#     with sync_playwright() as p:
#         browser = p.chromium.launch(
#             headless=True,
#             args=["--no-sandbox"]
#         )
#         page = browser.new_page()
#         page.context.add_cookies([
#             {
#                 "name": "session",
#                 "value": request.cookies.get("session"),
#                 "domain": request.host.split(":")[0],
#                 "path": "/"
#             }
#         ])

#         page.goto(
#             url,
#             wait_until='networkidle'
#         )

#         page.emulate_media(media='print')
#         pdf = page.pdf(
#             format='A4',
#             print_background=True,
#             scale=0.85,
#             margin={
#                 "top": "20px",
#                 "bottom": "20px",
#                 "left": "20px",
#                 "right": "20px"
#             }
#         )
#         browser.close()

#     return Response(
#         pdf,
#         mimetype='application/pdf',
#         headers={
#             'Content-Disposition':
#             f'attachment; filename="{customer_name}_analysis.pdf"'
#         }
#     )

@app.route('/download-customer-pdf/<customer_name>')
def download_customer_pdf(customer_name):

    import io
    import pandas as pd

    from flask import send_file
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import A4

    from database import engine

    # =====================================
    # LOAD DATA
    # =====================================

    query = """
        SELECT *
        FROM history_forecast
        WHERE customer = %(customer)s
        ORDER BY id DESC
        LIMIT 10
    """

    df = pd.read_sql(
        query,
        engine,
        params={
            'customer': customer_name
        }
    )

    # =====================================
    # PDF BUFFER
    # =====================================

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []

    styles = getSampleStyleSheet()

    # =====================================
    # TITLE
    # =====================================

    title = Paragraph(
        f"<b>Customer Analysis Report</b><br/>{customer_name}",
        styles['Title']
    )

    elements.append(title)
    elements.append(Spacer(1, 20))

    # =====================================
    # SUMMARY
    # =====================================

    total_forecast = len(df)

    summary = Paragraph(
        f"""
        <b>Total Forecast Records:</b> {total_forecast}<br/>
        <b>Generated From:</b> PayRisk Analytics System
        """,
        styles['BodyText']
    )

    elements.append(summary)
    elements.append(Spacer(1, 20))

    # =====================================
    # TABLE
    # =====================================

    if len(df) > 0:

        selected_cols = []

        for col in [
            'invoice_number',
            'prediction',
            'risk_level',
            'payment_delay_days'
        ]:
            if col in df.columns:
                selected_cols.append(col)

        table_data = [selected_cols]

        for _, row in df.iterrows():

            row_data = []

            for col in selected_cols:
                row_data.append(str(row[col]))

            table_data.append(row_data)

        table = Table(table_data)

        table.setStyle(TableStyle([

            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3A5F')),

            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),

            ('GRID', (0, 0), (-1, -1), 1, colors.grey),

            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),

            ('FONTSIZE', (0, 0), (-1, -1), 9),

        ]))

        elements.append(table)

    else:

        elements.append(
            Paragraph(
                "No forecast history available.",
                styles['BodyText']
            )
        )

    # =====================================
    # BUILD PDF
    # =====================================

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{customer_name}_analysis.pdf',
        mimetype='application/pdf'
    )

# =========================================
# RUN APP
# =========================================
if __name__ == '__main__':
    app.run(debug=True)